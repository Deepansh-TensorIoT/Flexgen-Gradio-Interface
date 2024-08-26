# main.py
import os
import time
from openpyxl import load_workbook
from excel_extraction.Utils.register_utils import (
    extract_table, 
    identify_header_row, 
    table_cleanup, 
    clean_sheet_name, 
    unmerge_and_fill, 
    drop_hidden_columns, 
    remove_numerical_top_rows, 
    get_text_above_header, 
    drop_hidden_rows,
    add_text_above_header_as_columns,
    drop_rows_with_same_values
)
from excel_extraction.Utils.bit_utils import clean_and_save_tables, separate_tables

from S3_utils import save_to_s3

def process_excel_sheet(file_path, dir_path, last_rows, vendor_name):
    try:
        # Load workbook
        count = 0
        workbook = load_workbook(filename=file_path, data_only=True)
        hidden_row_count = []
        # Unmerge cells and fill with values
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            if ws.sheet_state == 'hidden':
                continue
            
            unmerge_and_fill(ws)
            drop_hidden_columns(ws)

        # Save the workbook
        unmerged_file_path = file_path.replace('.xlsx', '_unmerged.xlsx')
        workbook.save(unmerged_file_path)
#         print(f"Unmerged file saved as {unmerged_file_path}")

        # Reload the evaluated workbook
        evaluated_wb = load_workbook(filename=unmerged_file_path)
        for sheet_name in evaluated_wb.sheetnames:
            header_row_index = 0
            ws = evaluated_wb[sheet_name]
            if ws.sheet_state == 'hidden':
#                 print(f"Skipping hidden sheet: {sheet_name}")
                continue
            
            # Process the visible sheet
            print(f"Processing visible sheet: {sheet_name}")
            header_row_index = identify_header_row(ws)
            if header_row_index is None:
                header_row_index = 0
                
            if header_row_index is not None:
                
                print(f"Identified header row in sheet '{sheet_name}' at row {header_row_index + 1}")
                last_row_first_table = last_rows[count]
                if last_row_first_table.lower() == "no" or last_row_first_table.lower() == "No" or last_row_first_table == '':
                    count += 1
                    continue
             
                last_row_first_table = int(last_row_first_table) 
                print("Last row after deducting hidden rows: ", last_row_first_table)
                
                has_register_table = last_row_first_table != 0
                if has_register_table:
                    if header_row_index!=0:
                        Text_above_header = get_text_above_header(ws, header_row_index)
                    else:
                        Text_above_header = ''
                    register_table = extract_table(ws, header_row_index + 1, last_row_first_table)
                    register_table_cleaned = table_cleanup(register_table)
                    register_table_cleaned = remove_numerical_top_rows(register_table_cleaned)
                    register_table_cleaned = register_table_cleaned.drop_duplicates()
                    register_table_cleaned = drop_rows_with_same_values(register_table_cleaned)
                    if Text_above_header!='':
                        register_table_cleaned = add_text_above_header_as_columns(register_table_cleaned, Text_above_header)
                    register_dir = os.path.join(dir_path, "Register")
                    os.makedirs(register_dir, exist_ok=True)
                    register_csv_path = os.path.join(register_dir, f"{sheet_name}.csv")
                    register_table_cleaned.to_csv(register_csv_path, index=False, header=False)
                    print(f"Saved RegisterTable_{sheet_name}.csv in {register_dir}")

#                 print("Processing Bit Tables")
                
                tables = separate_tables(ws, last_row_first_table)
                clean_and_save_tables(tables, sheet_name, dir_path)
                count += 1
            else:
                
                count += 1
#                 print(f"No identifiable table found in sheet: {sheet_name}")

    except Exception as e:
        print(f"Error processing {file_path}: {e}")


def process_excel_files(temp_folder_path, vendor_path, last_rows, vendor_name):
    base_dir_path = os.path.join(temp_folder_path, f'{vendor_name}/csv')
    os.makedirs(os.path.join(base_dir_path, 'Register'), exist_ok=True)
    os.makedirs(os.path.join(base_dir_path, 'Bit'), exist_ok=True)
    print(f"Processing File: {vendor_path}")
    process_excel_sheet(vendor_path, base_dir_path, last_rows, vendor_name)
    print(f"Processed and saved files in {base_dir_path}.")
    print('*' * 120)
    
    return base_dir_path


def run_extraction(vendor_name,vendor_path, last_rows, model_input, version_input):
    # Example usage
    start = time.time()
    #folder_path = os.path.join('Excel_folder', vendor_name)
    #os.makedirs(folder_path, exist_ok=True)
    temp_folder_path = os.path.dirname(vendor_path)
    # Process Excel files
    last_rows_list = [row.strip() for row in last_rows.strip('[]').split(',')]
    print(last_rows_list)
    csv_dir_path = process_excel_files(temp_folder_path,vendor_path, last_rows_list, vendor_name)
    print("Process completed. Cleaned CSV files stored in:", csv_dir_path)
    end = time.time()
    print('-' * 120)
    #print("Total excel files processed:", len(find_excel_files(vendor_path)))
    print("Time taken to process files:", end - start)
    save_to_s3(vendor_name, model_input, version_input, csv_dir_path, folder_name = "intermediate_output_csvs")
    return csv_dir_path
