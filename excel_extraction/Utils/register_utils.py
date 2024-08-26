# register_utils.py
import pandas as pd
from openpyxl import load_workbook, Workbook
from os import path 
import re
import time
from datetime import datetime


def drop_hidden_rows(ws):
    rows_to_delete = []
    first_hidden_row = None
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        if ws.row_dimensions[row[0].row].hidden:
            rows_to_delete.append(row[0].row)
            count = count + 1
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
        
    if len(rows_to_delete) > 0:
        first_hidden_row  = rows_to_delete[0]
    return count, first_hidden_row
        
def has_at_least_three_non_empty_cells(row):
    non_empty_cells = [cell for cell in row if cell is not None and str(cell).lower() != 'nan']
    return len(non_empty_cells) >= 3

def is_non_numerical_header(row):
    for cell in row:
        if cell is not None and not isinstance(cell, (int, float)):
            return True
    return False

def has_mostly_unique_values(row):
    values = [str(cell).strip().lower() for cell in row if cell is not None and str(cell).lower() != 'nan']
    unique_values = set(values)
    return sum(values.count(value) for value in unique_values) - len(unique_values) <= 2

def is_mostly_english(text):
    def is_english(s):
        try:
            s.encode('ascii')
        except UnicodeEncodeError:
            return False
        else:
            return True

    if isinstance(text, str):
        english_chars = sum(1 for char in text if is_english(char))
        total_chars = len(text)
        return english_chars / total_chars > 0.7
    return True  # If not a string, consider it as valid

def cell_contains_entirely_numbers(text):
    if isinstance(text, (int, float)):
        return True
    if isinstance(text, str) and text.isdigit():
        return True
    return False

def cell_contains_mostly_numbers(text):
    if isinstance(text, (int, float)):
        return True
    
    if isinstance(text, str):
        num_digits = sum(char.isdigit() for char in text)
        total_chars = len(text)
        return (num_digits / total_chars) > 0.5
    
    return False


def is_valid_header_row(row):
    row = [cell for cell in row if cell is not None]
    non_empty_cells = [cell for cell in row if cell is not None and str(cell).strip().lower() != 'nan']
    if len(non_empty_cells) <= 3:
        return False

    num_non_numeric = sum(1 for cell in row if cell is not None and not isinstance(cell, (int, float)) and is_mostly_english(str(cell)))
    if num_non_numeric < len(row) / 2:
        return False

    for cell in row:
        if cell_contains_entirely_numbers(cell) or cell_contains_mostly_numbers(cell):
            #print(f"Row skipped (cell contains entirely numerical values indicating its part of data): {row}")
            return False
            
    values = [str(cell).strip().lower() for cell in row if cell is not None and str(cell).strip().lower() != 'nan']
    unique_values = set(values)
    if sum(values.count(value) for value in unique_values) - len(unique_values) >= 3:
        return False

    return True

def is_numerical_row(row):
    for cell in row:
        if cell is None or cell == '':
            continue
        if isinstance(cell, datetime):
            continue
        try:
            float(cell)
        except ValueError:
            return False
    return True


def is_mostly_chinese(row):
    def is_chinese(s):
        try:
            s.encode('ascii')
            return False
        except UnicodeEncodeError:
            return True

    chinese_char_count = sum(len(str(cell)) for cell in row if is_chinese(str(cell)))
    total_char_count = sum(len(str(cell)) for cell in row)
    return chinese_char_count / total_char_count > 0.7 if total_char_count > 0 else False

def drop_empty_columns(ws):
    empty_columns = []
    for col in ws.iter_cols(values_only=True):
        if all(cell is None for cell in col):
            empty_columns.append(ws.max_column)
            ws.delete_cols(ws.max_column)
    return ws

def drop_hidden_columns(ws):
    cols_to_delete = []
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, values_only=False):
        if ws.column_dimensions[col[0].column_letter].hidden:
            cols_to_delete.append(col[0].column)

    for col in reversed(cols_to_delete):
        ws.delete_cols(col)

def identify_header_row(ws):
    latest_header_row_index = None
    for row_index, row in enumerate(ws.iter_rows(values_only=True, max_row=10), start=1):
        if ws.row_dimensions[row_index].hidden:
            continue
        if is_numerical_row(row):
            continue
        if is_valid_header_row(row):
            latest_header_row_index = row_index - 1
            print(f"Updated potential header row to index: {latest_header_row_index}")
    
    if latest_header_row_index is not None:
            print(" ")
            print(f"Identified header row at index: {latest_header_row_index + 1}")
    return latest_header_row_index

# def extract_table(ws, start_row, end_row):
#     data = []
#     for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
#         if ws.row_dimensions[row[0]].hidden: 
#             pass
#         else:
#             data.append([cell for cell in row])
        
#     return pd.DataFrame(data)

def extract_table(ws, start_row, end_row): 
    data = [] 
    for row_index, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True), start=start_row): 
        if not ws.row_dimensions[row_index].hidden: 
            data.append([cell for cell in row]) 
    return pd.DataFrame(data)

def clean_sheet_name(sheet_name):
    return ''.join(char for char in sheet_name if ord(char) < 128)

def contains_non_english_numerical(text):
    valid_characters = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789/[]-_(){}!@#$%^&*+=|\\;:'\",.<>?~` °℃Ω")
    
    if isinstance(text, str):
        non_english_numerical_chars = sum(1 for char in text if char not in valid_characters)
        total_chars = len(text)
        return non_english_numerical_chars / total_chars > 0.3 if total_chars > 0 else False
    return False

def should_drop_based_on_non_english_content(df, column_name):
    non_english_content_count = df[column_name].apply(contains_non_english_numerical).sum()
    total_elements = len(df[column_name])
    return non_english_content_count / total_elements > 0.3 if total_elements > 0 else False

def drop_columns_more_than_90_empty(df):
    threshold = 0.9
    num_rows = len(df)
    columns_to_drop = [col for col in df.columns if df[col].isna().sum() / num_rows > threshold]
    df = df.drop(columns=columns_to_drop)
    return df

def table_cleanup(df):
    df.columns = df.columns.astype(str)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    df = df.dropna(axis=1, how='all').dropna(axis=0, how='all')
    df = drop_columns_more_than_90_empty(df)
    columns_to_drop = [col for col in df.columns if should_drop_based_on_non_english_content(df, col)]
    #print("Columns to drop: ", columns_to_drop)
    df = df.drop(columns=columns_to_drop)
    return df

def unmerge_and_fill(ws):
    for m_range in list(ws.merged_cells.ranges):
        merged_cell = m_range.start_cell
        merged_value = merged_cell.value
        ws.unmerge_cells(range_string=str(m_range))
        for row in range(m_range.min_row, m_range.max_row + 1):
            for col in range(m_range.min_col, m_range.max_col + 1):
                ws.cell(row=row, column=col, value=merged_value)

def remove_numerical_top_rows(df):
    while not df.empty and is_numerical_row(df.iloc[0]):
        df = df.iloc[1:]
    return df

def drop_rows_with_same_values(df): # Function to check if all values in a row are the same
    
    def all_same(row): return all(x == row[0] for x in row) 
    
    # Apply the function to each row and drop the rows that meet the condition 
    df = df[~df.apply(all_same, axis=1)]
    return df
    

def get_text_above_header(ws, header_row_index):
    text_above_header = []
    header_row_values = [cell.value for cell in ws[header_row_index + 1]]
    for row in ws.iter_rows(min_row=1, max_row=header_row_index, values_only=True):
        row_text = []
        for cell in row:
            if cell is not None and cell not in header_row_values:
                row_text.append(str(cell))
        if row_text:
            text_above_header.append(' '.join(row_text))
    
    # Flatten the list and remove duplicates
    unique_text_above_header = ' '.join(dict.fromkeys(' '.join(text_above_header).split()))
    return unique_text_above_header

def is_mostly_english_header(s, letter_threshold=0.9):
    letters = sum(c.isalpha() for c in s)
    digits = sum(c.isdigit() for c in s)
    total = len(s)

    # Calculate the proportions
    letter_ratio = letters / total if total > 0 else 0
    digit_ratio = digits / total if total > 0 else 0

    # Check if the letter ratio meets the threshold and the digit ratio is low
    return letter_ratio >= letter_threshold and digit_ratio <= (1 - letter_threshold)

def add_text_above_header_as_columns(df, text_above_header):
    if text_above_header and is_mostly_english_header(text_above_header):
        text_column = ['Text Above Header'] + [text_above_header] * (len(df) - 1)
        df.insert(len(df.columns), 'Text Above Header', text_column)
    return df